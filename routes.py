from flask import render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
from app import app, db
from models import Learner, LevelTest, Counseling, Exam, ConsultationRecord, LearnerHistory, PrivacyConsent
from forms import LearnerApplicationForm, CounselingForm, ExamForm, UploadForm, MemoForm, AcademicStatusForm, ClassAssignmentForm, ConsultationRecordForm, LearnerHistoryForm, PrivacyConsentForm
from datetime import datetime, date
from sqlalchemy import func, extract
import json
import pandas as pd
import os
from werkzeug.utils import secure_filename
import io
import openpyxl
import tempfile
from ai_analysis import transcribe_audio, summarize_counseling_session

# 레벨 테스트 문제 데이터 (원본 문항)
LEVEL_TEST_QUESTIONS = {
    'basic': [
        {
            'id': 1,
            'type': 'multiple_choice',
            'question': '다음 중 자음이 아닌 것은?',
            'options': ['ㄱ', 'ㅏ', 'ㄴ', 'ㄷ'],
            'correct': 1,
            'category': 'vocabulary'
        },
        {
            'id': 2,
            'type': 'multiple_choice',
            'question': '"ㄱ + ㅏ"를 합치면?',
            'options': ['가', '고', '구', '기'],
            'correct': 0,
            'category': 'vocabulary'
        },
        {
            'id': 3,
            'type': 'multiple_choice',
            'question': '다음 중 받침이 있는 글자는?',
            'options': ['가', '강', '고', '구'],
            'correct': 1,
            'category': 'grammar'
        },
        {
            'id': 4,
            'type': 'drawing',
            'question': '다음 글자를 써보세요: 가',
            'category': 'grammar'
        },
        {
            'id': 5,
            'type': 'drawing',
            'question': '다음 글자를 써보세요: 나',
            'category': 'grammar'
        }
    ],
    'intermediate': [
        {
            'id': 6,
            'type': 'multiple_choice',
            'question': '다음 단어를 올바르게 읽은 것은? 우리',
            'options': ['우리', '오리', '울이', '으리'],
            'correct': 0,
            'category': 'reading'
        },
        {
            'id': 7,
            'type': 'multiple_choice',
            'question': '받침이 있는 단어를 올바르게 읽은 것은? 웃다',
            'options': ['우다', '웃다', '욷다', '웃따'],
            'correct': 1,
            'category': 'reading'
        },
        {
            'id': 8,
            'type': 'drawing',
            'question': '다음 단어를 써보세요: 집',
            'category': 'grammar'
        }
    ],
    'advanced': [
        {
            'id': 9,
            'type': 'reading',
            'question': '다음 글을 읽고 질문에 답하세요.\n"오늘은 날씨가 좋다. 하늘이 맑고 바람이 시원하다."\n이 글에서 오늘 날씨는 어떻습니까?',
            'options': ['좋다', '나쁘다', '추웠다', '더웠다'],
            'correct': 0,
            'category': 'reading'
        },
        {
            'id': 10,
            'type': 'drawing',
            'question': '자신의 이름을 써보세요.',
            'category': 'grammar'
        }
    ]
}

@app.route('/')
def index():
    """메인 페이지"""
    # 통계 데이터 계산
    total_learners = Learner.query.count()
    completed_tests = LevelTest.query.count()
    
    # 반별 분포
    class_distribution = db.session.query(
        Learner.assigned_class, 
        func.count(Learner.id)
    ).filter(Learner.assigned_class.isnot(None)).group_by(Learner.assigned_class).all()
    
    # 최근 등록자
    recent_learners = Learner.query.order_by(Learner.created_at.desc()).limit(5).all()
    
    return render_template('index.html',
                         total_learners=total_learners,
                         completed_tests=completed_tests,
                         class_distribution=class_distribution,
                         recent_learners=recent_learners)

@app.route('/application', methods=['GET', 'POST'])
def application():
    """학습자 지원서 작성"""
    form = LearnerApplicationForm()
    
    if form.validate_on_submit():
        learner = Learner(
            # 기본 인적사항
            name=form.name.data,
            birth_date=form.birth_date.data,
            gender=form.gender.data,
            phone=form.phone.data,
            address=form.address.data,
            
            # 비상연락처
            emergency_contact1_name=form.emergency_contact1_name.data,
            emergency_contact1_relation=form.emergency_contact1_relation.data,
            emergency_contact1_phone=form.emergency_contact1_phone.data,
            emergency_contact2_name=form.emergency_contact2_name.data,
            emergency_contact2_relation=form.emergency_contact2_relation.data,
            emergency_contact2_phone=form.emergency_contact2_phone.data,
            
            # 학습 배경 및 건강 상태
            health_status=form.health_status.data,
            health_details=form.health_details.data,
            surgery_experience=form.surgery_experience.data,
            surgery_details=form.surgery_details.data,
            learning_experience=form.learning_experience.data,
            learning_institution=form.learning_institution.data,
            learning_period=form.learning_period.data,
            
            # 입학 경로 및 학습 목적
            enrollment_path=form.enrollment_path.data,
            learning_purpose=form.learning_purpose.data,
            
            # 기초 진단 평가
            reading_level=form.reading_level.data,
            writing_level=form.writing_level.data,
            math_level=form.math_level.data,
            
            # 추가 정보
            other_info=form.other_info.data,
            
            # 개인정보 동의 및 서명
            privacy_agreement=form.privacy_consent.data,
            digital_signature=form.signature_data.data,
            
            status='지원서 작성'
        )
        
        db.session.add(learner)
        db.session.commit()
        
        # AI 분석 생성
        try:
            from ai_analysis import analyze_learner_profile, format_analysis_for_display
            
            # 학습자 데이터를 딕셔너리로 변환
            learner_data = {
                'name': learner.name,
                'birth_date': learner.birth_date.strftime('%Y-%m-%d') if learner.birth_date else '',
                'gender': learner.gender,
                'phone': learner.phone,
                'health_status': learner.health_status,
                'health_details': learner.health_details,
                'surgery_experience': learner.surgery_experience,
                'surgery_details': learner.surgery_details,
                'learning_experience': learner.learning_experience,
                'learning_institution': learner.learning_institution,
                'learning_period': learner.learning_period,
                'enrollment_path': learner.enrollment_path,
                'learning_purpose': learner.learning_purpose,
                'reading_level': learner.reading_level,
                'writing_level': learner.writing_level,
                'math_level': learner.math_level,
                'other_info': learner.other_info
            }
            
            # AI 분석 실행
            analysis_result = analyze_learner_profile(learner_data)
            learner.ai_analysis = format_analysis_for_display(analysis_result)
            db.session.commit()
            
        except Exception as e:
            # AI 분석 실패 시 로그만 남기고 계속 진행
            print(f"AI 분석 실패: {e}")
            learner.ai_analysis = "AI 분석을 수행할 수 없습니다. 수동으로 검토해주세요."
            db.session.commit()
        
        # 세션에 학습자 ID 저장
        session['learner_id'] = learner.id
        session['workflow_step'] = 'level_test'
        
        flash('지원서가 성공적으로 제출되었습니다. 레벨 테스트를 진행해주세요.', 'success')
        return redirect(url_for('level_test'))
    
    return render_template('application.html', form=form)

@app.route('/level_test')
def level_test():
    """레벨 테스트 페이지"""
    # URL 파라미터에서 learner_id 확인 (기존 학습자가 테스트 응시하는 경우)
    learner_id_param = request.args.get('learner_id')
    
    if learner_id_param:
        # 기존 학습자가 레벨 테스트를 응시하는 경우
        learner_id = learner_id_param
        session['learner_id'] = learner_id  # 세션에 저장
        is_retake = True
    elif 'learner_id' in session:
        # 신규 지원자가 지원서 작성 후 레벨 테스트를 응시하는 경우
        learner_id = session['learner_id']
        is_retake = False
    else:
        flash('학습자 정보를 찾을 수 없습니다.', 'warning')
        return redirect(url_for('learner_list'))
    
    learner = Learner.query.get_or_404(learner_id)
    
    # 동적 문제 로딩으로 변경
    dynamic_questions = {
        'basic': get_level_test_questions('basic'),
        'intermediate': get_level_test_questions('intermediate'), 
        'advanced': get_level_test_questions('advanced')
    }
    
    return render_template('level_test.html', 
                         learner=learner, 
                         questions=dynamic_questions,
                         is_retake=is_retake)

@app.route('/submit_level_test', methods=['POST'])
def submit_level_test():
    """레벨 테스트 결과 제출"""
    if 'learner_id' not in session:
        return jsonify({'error': '세션이 만료되었습니다.'}), 400
    
    learner = Learner.query.get(session['learner_id'])
    if not learner:
        return jsonify({'error': '학습자 정보를 찾을 수 없습니다.'}), 400
    
    data = request.get_json()
    
    # 점수 계산
    vocabulary_score = 0
    grammar_score = 0
    reading_score = 0
    
    # 각 영역별 점수 계산 (기초/중급/고급 단계별로)
    all_answers = {**data.get('basic_answers', {}), 
                   **data.get('intermediate_answers', {}), 
                   **data.get('advanced_answers', {})}
    
    for question_id, answer in all_answers.items():
        q_id = int(question_id)
        # 문제 찾기
        question = None
        for stage in LEVEL_TEST_QUESTIONS.values():
            for q in stage:
                if q['id'] == q_id:
                    question = q
                    break
        
        if question and question['type'] == 'multiple_choice':
            if answer == question['correct']:
                score = 20 if q_id <= 5 else (25 if q_id <= 8 else 30)  # 기초:20점, 중급:25점, 고급:30점
                if question['category'] == 'vocabulary':
                    vocabulary_score += score
                elif question['category'] == 'grammar':
                    grammar_score += score
                elif question['category'] == 'reading':
                    reading_score += score
    
    total_score = vocabulary_score + grammar_score + reading_score
    
    # 레벨 테스트 결과 저장
    level_test = LevelTest(
        learner_id=learner.id,
        vocabulary_score=vocabulary_score,
        grammar_score=grammar_score,
        reading_score=reading_score,
        total_score=total_score,
        basic_answers=data.get('basic_answers', {}),
        intermediate_answers=data.get('intermediate_answers', {}),
        advanced_answers=data.get('advanced_answers', {}),
        drawing_data=data.get('drawing_data', '')
    )
    
    db.session.add(level_test)
    
    # 학습자 상태 업데이트
    learner.status = '레벨 테스트'
    db.session.commit()
    
    session['workflow_step'] = 'counseling'
    
    # 기존 학습자인지 신규 학습자인지 확인하여 적절한 페이지로 리다이렉트
    is_retake = bool(request.args.get('learner_id') or session.get('is_retake', False))
    
    if is_retake:
        # 기존 학습자의 경우 학습자 프로필로 돌아가기
        redirect_url = url_for('learner_profile', learner_id=learner.id)
    else:
        # 신규 학습자의 경우 상담 페이지로 이동
        redirect_url = url_for('counseling')
    
    return jsonify({
        'success': True,
        'total_score': total_score,
        'vocabulary_score': vocabulary_score,
        'grammar_score': grammar_score,
        'reading_score': reading_score,
        'redirect_url': redirect_url
    })

@app.route('/counseling', methods=['GET', 'POST'])
def counseling():
    """상담 페이지"""
    if 'learner_id' not in session:
        flash('먼저 지원서를 작성해주세요.', 'warning')
        return redirect(url_for('application'))
    
    learner = Learner.query.get(session['learner_id'])
    if not learner:
        flash('학습자 정보를 찾을 수 없습니다.', 'error')
        return redirect(url_for('application'))
    
    # 레벨 테스트 결과 확인
    level_test = LevelTest.query.filter_by(learner_id=learner.id).first()
    
    form = CounselingForm()
    
    if form.validate_on_submit():
        counseling = Counseling(
            learner_id=learner.id,
            counselor_name=form.counselor_name.data,
            counseling_content=form.counseling_content.data,
            recommendations=form.recommendations.data,
            notes=form.notes.data
        )
        
        db.session.add(counseling)
        
        # 반 배정 로직
        if level_test:
            if level_test.total_score >= 80:
                assigned_class = '예비 중등반'
            elif level_test.total_score >= 60:
                assigned_class = '초등 고급반'
            elif level_test.total_score >= 40:
                assigned_class = '초등 중급반'
            else:
                assigned_class = '초등 초급반'
            
            learner.assigned_class = assigned_class
        
        learner.status = '반 배정 완료'
        db.session.commit()
        
        session['workflow_step'] = 'completed'
        
        flash('상담이 완료되었습니다. 반 배정 결과를 확인해주세요.', 'success')
        return redirect(url_for('assignment_result'))
    
    return render_template('counseling.html', form=form, learner=learner, level_test=level_test)

@app.route('/assignment_result')
def assignment_result():
    """반 배정 결과 페이지"""
    if 'learner_id' not in session:
        flash('먼저 지원서를 작성해주세요.', 'warning')
        return redirect(url_for('application'))
    
    learner = Learner.query.get(session['learner_id'])
    if not learner:
        flash('학습자 정보를 찾을 수 없습니다.', 'error')
        return redirect(url_for('application'))
    
    level_test = LevelTest.query.filter_by(learner_id=learner.id).first()
    counseling = Counseling.query.filter_by(learner_id=learner.id).first()
    
    # 세션 정리
    session.pop('learner_id', None)
    session.pop('workflow_step', None)
    
    return render_template('assignment_result.html', 
                         learner=learner, 
                         level_test=level_test, 
                         counseling=counseling)

@app.route('/learners')
def learner_list():
    """학습자 목록 페이지"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    class_filter = request.args.get('class_filter', '')
    actual_class_filter = request.args.get('actual_class_filter', '')
    status_filter = request.args.get('status_filter', '')
    
    query = Learner.query
    
    if search:
        query = query.filter(Learner.name.contains(search))
    
    if class_filter:
        query = query.filter(Learner.assigned_class == class_filter)
    
    if actual_class_filter:
        query = query.filter(Learner.class_name == actual_class_filter)
    
    if status_filter:
        query = query.filter(Learner.status == status_filter)
    
    learners = query.order_by(Learner.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # 필터 옵션
    classes = db.session.query(Learner.assigned_class).filter(
        Learner.assigned_class.isnot(None)
    ).distinct().all()
    actual_classes = db.session.query(Learner.class_name).filter(
        Learner.class_name.isnot(None)
    ).distinct().all()
    statuses = db.session.query(Learner.status).distinct().all()
    
    return render_template('learner_list.html', 
                         learners=learners,
                         search=search,
                         class_filter=class_filter,
                         actual_class_filter=actual_class_filter,
                         status_filter=status_filter,
                         classes=[c[0] for c in classes],
                         actual_classes=[c[0] for c in actual_classes],
                         statuses=[s[0] for s in statuses])

@app.route('/learner/<int:learner_id>')
def learner_profile(learner_id):
    """학습자 상세 프로필"""
    learner = Learner.query.get_or_404(learner_id)
    level_tests = LevelTest.query.filter_by(learner_id=learner_id).all()
    counselings = Counseling.query.filter_by(learner_id=learner_id).order_by(Counseling.created_at.desc()).all()
    exams = Exam.query.filter_by(learner_id=learner_id).order_by(Exam.exam_date.desc()).all()
    
    return render_template('learner_profile.html',
                         learner=learner,
                         level_tests=level_tests,
                         counselings=counselings,
                         exams=exams)

@app.route('/learner/<int:learner_id>/counseling', methods=['GET', 'POST'])
def add_counseling(learner_id):
    """추가 상담 기록"""
    learner = Learner.query.get_or_404(learner_id)
    form = CounselingForm()
    
    if form.validate_on_submit():
        counseling = Counseling(
            learner_id=learner.id,
            counselor_name=form.counselor_name.data,
            counseling_content=form.counseling_content.data,
            recommendations=form.recommendations.data,
            notes=form.notes.data
        )
        
        db.session.add(counseling)
        db.session.commit()
        
        flash('상담 기록이 추가되었습니다.', 'success')
        return redirect(url_for('learner_profile', learner_id=learner.id))
    
    return render_template('counseling.html', form=form, learner=learner, is_additional=True)

@app.route('/learner/<int:learner_id>/exam', methods=['GET', 'POST'])
def add_exam(learner_id):
    """시험 기록 추가"""
    learner = Learner.query.get_or_404(learner_id)
    form = ExamForm()
    
    if form.validate_on_submit():
        exam = Exam(
            learner_id=learner.id,
            exam_name=form.exam_name.data,
            exam_date=form.exam_date.data,
            score=form.score.data,
            max_score=form.max_score.data,
            notes=form.notes.data
        )
        
        db.session.add(exam)
        db.session.commit()
        
        flash('시험 기록이 추가되었습니다.', 'success')
        return redirect(url_for('learner_profile', learner_id=learner.id))
    
    return render_template('add_exam.html', form=form, learner=learner)

@app.route('/dashboard')
def dashboard():
    """통계 대시보드"""
    # 전체 통계
    total_learners = Learner.query.count()
    completed_tests = LevelTest.query.count()
    
    # 반별 분포
    class_distribution = db.session.query(
        Learner.assigned_class, 
        func.count(Learner.id)
    ).filter(Learner.assigned_class.isnot(None)).group_by(Learner.assigned_class).all()
    
    # 점수 분포
    score_ranges = [
        ('80점 이상', LevelTest.query.filter(LevelTest.total_score >= 80).count()),
        ('60-79점', LevelTest.query.filter(LevelTest.total_score.between(60, 79)).count()),
        ('40-59점', LevelTest.query.filter(LevelTest.total_score.between(40, 59)).count()),
        ('40점 미만', LevelTest.query.filter(LevelTest.total_score < 40).count())
    ]
    
    # 월별 등록 추이 (최근 6개월)
    monthly_registrations = db.session.query(
        extract('year', Learner.created_at).label('year'),
        extract('month', Learner.created_at).label('month'),
        func.count(Learner.id).label('count')
    ).group_by('year', 'month').order_by('year', 'month').limit(6).all()
    
    return render_template('dashboard.html',
                         total_learners=total_learners,
                         completed_tests=completed_tests,
                         class_distribution=class_distribution,
                         score_ranges=score_ranges,
                         monthly_registrations=monthly_registrations)

@app.route('/api/dashboard_data')
def dashboard_data():
    """대시보드 차트 데이터 API"""
    # 반별 분포 데이터
    class_data = db.session.query(
        Learner.assigned_class, 
        func.count(Learner.id)
    ).filter(Learner.assigned_class.isnot(None)).group_by(Learner.assigned_class).all()
    
    # 점수 분포 데이터
    score_data = [
        LevelTest.query.filter(LevelTest.total_score >= 80).count(),
        LevelTest.query.filter(LevelTest.total_score.between(60, 79)).count(),
        LevelTest.query.filter(LevelTest.total_score.between(40, 59)).count(),
        LevelTest.query.filter(LevelTest.total_score < 40).count()
    ]
    
    # 월별 등록 추이 데이터
    monthly_data = db.session.query(
        extract('year', Learner.created_at).label('year'),
        extract('month', Learner.created_at).label('month'),
        func.count(Learner.id).label('count')
    ).group_by('year', 'month').order_by('year', 'month').limit(12).all()
    
    # 실제 운영반별 분포 데이터
    actual_class_data = db.session.query(
        Learner.class_name, 
        func.count(Learner.id)
    ).filter(Learner.class_name.isnot(None)).group_by(Learner.class_name).all()
    
    return jsonify({
        'class_distribution': {
            'labels': [item[0] for item in class_data],
            'data': [item[1] for item in class_data]
        },
        'actual_class_distribution': {
            'labels': [item[0] for item in actual_class_data],
            'data': [item[1] for item in actual_class_data]
        },
        'score_distribution': {
            'labels': ['80점 이상', '60-79점', '40-59점', '40점 미만'],
            'data': score_data
        },
        'monthly_registrations': {
            'labels': [f"{int(item[0])}년 {int(item[1])}월" for item in monthly_data],
            'data': [item[2] for item in monthly_data]
        }
    })

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    """엑셀/CSV 업로드 페이지"""
    form = UploadForm()
    
    if form.validate_on_submit():
        file = form.file.data
        filename = secure_filename(file.filename)
        
        try:
            # 파일 확장자에 따라 읽기
            if filename.endswith('.csv'):
                df = pd.read_csv(file)
            else:  # Excel files
                df = pd.read_excel(file)
            
            # 필수 컬럼 확인
            required_columns = ['이름', '연락처', '이메일', '배정반', '상태']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                flash(f'필수 컬럼이 누락되었습니다: {", ".join(missing_columns)}', 'error')
                return render_template('upload.html', form=form)
            
            # 데이터 처리 및 저장
            success_count = 0
            error_count = 0
            
            for index, row in df.iterrows():
                try:
                    # 중복 확인 (이름과 연락처로)
                    existing = Learner.query.filter_by(
                        name=row['이름'],
                        phone=row['연락처']
                    ).first()
                    
                    if existing:
                        # 기존 데이터 업데이트
                        existing.assigned_class = row['배정반'] if pd.notna(row['배정반']) else None
                        existing.status = row['상태'] if pd.notna(row['상태']) else '지원서 작성'
                        success_count += 1
                    else:
                        # 새 학습자 생성
                        learner = Learner(
                            name=row['이름'],
                            phone=row['연락처'],
                            birth_date=date.today(),  # 기본값
                            gender='남',  # 기본값
                            address='업로드 데이터',  # 기본값
                            emergency_contact1_name='미입력',  # 기본값
                            emergency_contact1_relation='미입력',  # 기본값
                            emergency_contact1_phone='미입력',  # 기본값
                            health_status='건강',  # 기본값
                            surgery_experience='없음',  # 기본값
                            learning_experience='없음',  # 기본값
                            enrollment_path='기타',  # 기본값
                            learning_purpose='기타',  # 기본값
                            reading_level='전혀 못 읽음',  # 기본값
                            writing_level='전혀 못 씀',  # 기본값
                            math_level='숫자 전혀 모름',  # 기본값
                            assigned_class=row['배정반'] if pd.notna(row['배정반']) else None,
                            status=row['상태'] if pd.notna(row['상태']) else '지원서 작성'
                        )
                        db.session.add(learner)
                        success_count += 1
                        
                except Exception as e:
                    error_count += 1
                    print(f"Row {index} error: {e}")
                    continue
            
            db.session.commit()
            flash(f'업로드 완료: {success_count}명 처리, {error_count}명 오류', 'success')
            return redirect(url_for('learner_list'))
            
        except Exception as e:
            flash(f'파일 처리 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return render_template('upload.html', form=form)

@app.route('/download_template')
def download_template():
    """업로드 템플릿 다운로드"""
    # 템플릿 데이터 생성
    template_data = {
        '이름': ['홍길동', '김철수', '이영희'],
        '연락처': ['010-1234-5678', '010-2345-6789', '010-3456-7890'],
        '이메일': ['hong@example.com', 'kim@example.com', 'lee@example.com'],
        '배정반': ['초등 초급반', '초등 중급반', '초등 고급반'],
        '상태': ['지원서 작성', '레벨 테스트', '상담']
    }
    
    df = pd.DataFrame(template_data)
    
    # 메모리에서 Excel 파일 생성
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='학습자목록')
    
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=learner_template.xlsx'
    
    return response

@app.route('/edit_memo/<int:learner_id>', methods=['GET', 'POST'])
def edit_memo(learner_id):
    """학습자 메모 수정"""
    learner = Learner.query.get_or_404(learner_id)
    form = MemoForm()
    
    if form.validate_on_submit():
        learner.memo = form.memo.data
        db.session.commit()
        flash('메모가 저장되었습니다.', 'success')
        return redirect(url_for('learner_profile', learner_id=learner.id))
    
    # GET 요청 시 기존 메모 로드
    form.memo.data = learner.memo
    
    return render_template('edit_memo.html', form=form, learner=learner)

@app.route('/edit_academic_status/<int:learner_id>', methods=['GET', 'POST'])
def edit_academic_status(learner_id):
    """학적상태 변경"""
    learner = Learner.query.get_or_404(learner_id)
    form = AcademicStatusForm()
    
    if request.method == 'GET':
        form.academic_status.data = learner.academic_status or '재학'
    
    if form.validate_on_submit():
        learner.academic_status = form.academic_status.data
        db.session.commit()
        flash(f'{learner.name}의 학적상태가 {form.academic_status.data}(으)로 변경되었습니다.', 'success')
        return redirect(url_for('learner_profile', learner_id=learner.id))
    
    return render_template('edit_academic_status.html', form=form, learner=learner)

@app.route('/edit_class_assignment/<int:learner_id>', methods=['GET', 'POST'])
def edit_class_assignment(learner_id):
    """반 배정 변경"""
    learner = Learner.query.get_or_404(learner_id)
    form = ClassAssignmentForm()
    
    if request.method == 'GET':
        form.class_name.data = learner.class_name or ''
    
    if form.validate_on_submit():
        old_class = learner.class_name
        new_class = form.class_name.data if form.class_name.data else None
        learner.class_name = new_class
        
        # 이력 기록 생성
        if old_class != new_class:
            if new_class:
                event_title = f"{new_class} 배정"
                event_description = f"{old_class or '미배정'}에서 {new_class}로 반 변경"
            else:
                event_title = "반 배정 해제"
                event_description = f"{old_class}에서 미배정으로 변경"
            
            history_record = LearnerHistory(
                learner_id=learner.id,
                event_type='class_assignment',
                event_date=date.today(),
                event_title=event_title,
                event_description=event_description,
                previous_value=old_class,
                new_value=new_class,
                created_by='시스템'
            )
            db.session.add(history_record)
        
        db.session.commit()
        
        if new_class:
            flash(f'{learner.name}이(가) {new_class}에 배정되었습니다.', 'success')
        else:
            flash(f'{learner.name}의 반 배정이 해제되었습니다.', 'success')
        return redirect(url_for('learner_profile', learner_id=learner.id))
    
    return render_template('edit_class_assignment.html', form=form, learner=learner)

@app.route('/consultation_record', methods=['GET', 'POST'])
def consultation_record():
    """상담 기록 작성"""
    form = ConsultationRecordForm()
    
    if form.validate_on_submit():
        record = ConsultationRecord(
            name=form.name.data or None,
            phone=form.phone.data,
            consultation_content=form.consultation_content.data or None,
            consultation_date=form.consultation_date.data,
            gender=form.gender.data or None,
            residence=form.residence.data or None,
            consultation_area=form.consultation_area.data or None
        )
        
        db.session.add(record)
        db.session.commit()
        
        flash('상담 기록이 저장되었습니다.', 'success')
        return redirect(url_for('consultation_list'))
    
    return render_template('consultation_record.html', form=form)

@app.route('/consultation_list')
def consultation_list():
    """상담 기록 목록"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    
    query = db.session.query(ConsultationRecord.phone, 
                           func.max(ConsultationRecord.name).label('name'),
                           func.count(ConsultationRecord.id).label('count'),
                           func.max(ConsultationRecord.consultation_date).label('latest_date')
                          ).group_by(ConsultationRecord.phone)
    
    if search:
        query = query.filter(
            (ConsultationRecord.phone.contains(search)) | 
            (ConsultationRecord.name.contains(search))
        )
    
    results = query.order_by(func.max(ConsultationRecord.consultation_date).desc()).all()
    
    # 페이지네이션을 위한 수동 처리
    total = len(results)
    per_page = 20
    start = (page - 1) * per_page
    end = start + per_page
    
    paginated_results = results[start:end]
    
    # 간단한 페이지네이션 객체 만들기
    class SimplePagination:
        def __init__(self, items, total, page, per_page):
            self.items = items
            self.total = total
            self.page = page
            self.per_page = per_page
            self.pages = (total - 1) // per_page + 1 if total > 0 else 1
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if self.has_prev else None
            self.next_num = page + 1 if self.has_next else None
        
        def iter_pages(self):
            for p in range(max(1, self.page - 2), min(self.pages + 1, self.page + 3)):
                yield p
    
    consultation_groups = SimplePagination(paginated_results, total, page, per_page)
    
    return render_template('consultation_list.html', 
                         consultation_groups=consultation_groups,
                         search=search)

@app.route('/consultation_detail/<phone>')
def consultation_detail(phone):
    """특정 전화번호의 상담 기록 상세"""
    records = ConsultationRecord.query.filter_by(phone=phone).order_by(
        ConsultationRecord.consultation_date.desc()
    ).all()
    
    if not records:
        flash('해당 전화번호의 상담 기록을 찾을 수 없습니다.', 'error')
        return redirect(url_for('consultation_list'))
    
    return render_template('consultation_detail.html', records=records, phone=phone)

@app.route('/add_learner_history/<int:learner_id>', methods=['GET', 'POST'])
def add_learner_history(learner_id):
    """학습자 이력 추가"""
    learner = Learner.query.get_or_404(learner_id)
    form = LearnerHistoryForm()
    
    if form.validate_on_submit():
        history_record = LearnerHistory(
            learner_id=learner.id,
            event_type=form.event_type.data,
            event_date=form.event_date.data,
            event_title=form.event_title.data,
            event_description=form.event_description.data or None,
            previous_value=form.previous_value.data or None,
            new_value=form.new_value.data or None,
            created_by=form.created_by.data or '관리자'
        )
        
        db.session.add(history_record)
        db.session.commit()
        
        flash('이력이 추가되었습니다.', 'success')
        return redirect(url_for('learner_profile', learner_id=learner.id))
    
    return render_template('add_learner_history.html', form=form, learner=learner)

@app.route('/delete_learner_history/<int:history_id>')
def delete_learner_history(history_id):
    """학습자 이력 삭제"""
    history_record = LearnerHistory.query.get_or_404(history_id)
    learner_id = history_record.learner_id
    
    db.session.delete(history_record)
    db.session.commit()
    
    flash('이력이 삭제되었습니다.', 'success')
    return redirect(url_for('learner_profile', learner_id=learner_id))

@app.route('/regenerate_ai_analysis/<int:learner_id>')
def regenerate_ai_analysis(learner_id):
    """기존 학습자의 AI 분석 재생성"""
    learner = Learner.query.get_or_404(learner_id)
    
    try:
        from ai_analysis import analyze_learner_profile, format_analysis_for_display
        
        # 학습자 데이터를 딕셔너리로 변환
        learner_data = {
            'name': learner.name,
            'birth_date': learner.birth_date.strftime('%Y-%m-%d') if learner.birth_date else '',
            'gender': learner.gender,
            'phone': learner.phone,
            'health_status': learner.health_status,
            'health_details': learner.health_details,
            'surgery_experience': learner.surgery_experience,
            'surgery_details': learner.surgery_details,
            'learning_experience': learner.learning_experience,
            'learning_institution': learner.learning_institution,
            'learning_period': learner.learning_period,
            'enrollment_path': learner.enrollment_path,
            'learning_purpose': learner.learning_purpose,
            'reading_level': learner.reading_level,
            'writing_level': learner.writing_level,
            'math_level': learner.math_level,
            'other_info': learner.other_info
        }
        
        # AI 분석 실행
        analysis_result = analyze_learner_profile(learner_data)
        learner.ai_analysis = format_analysis_for_display(analysis_result)
        db.session.commit()
        
        flash('AI 분석이 성공적으로 생성되었습니다.', 'success')
        
    except Exception as e:
        flash(f'AI 분석 생성 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('learner_profile', learner_id=learner_id))

@app.route('/learners/delete/<int:learner_id>', methods=['POST'])
def delete_learner(learner_id):
    """학습자 삭제"""
    learner = Learner.query.get_or_404(learner_id)
    
    try:
        # 관련된 모든 데이터 삭제
        LevelTest.query.filter_by(learner_id=learner_id).delete()
        Counseling.query.filter_by(learner_id=learner_id).delete()
        Exam.query.filter_by(learner_id=learner_id).delete()
        
        # 학습자 삭제
        db.session.delete(learner)
        db.session.commit()
        
        flash(f'{learner.name} 학습자가 성공적으로 삭제되었습니다.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'학습자 삭제 중 오류가 발생했습니다: {str(e)}', 'error')
    
    return redirect(url_for('learner_list'))

@app.route('/privacy_consent/<int:learner_id>', methods=['GET', 'POST'])
def privacy_consent(learner_id):
    """개인정보 동의서 작성/수정"""
    learner = Learner.query.get_or_404(learner_id)
    
    # 기존 동의서가 있는지 확인
    existing_consent = PrivacyConsent.query.filter_by(learner_id=learner_id).first()
    
    form = PrivacyConsentForm()
    
    # 기본 동의서 내용 설정 (수정 가능)
    default_content = """개인정보 수집 및 이용에 관한 동의서

난곡 사랑의집에서는 성인문해교육 프로그램 운영을 위하여 다음과 같이 개인정보를 수집하고 이용하고자 합니다.

1. 개인정보의 수집·이용 목적
   - 성인문해교육 프로그램 제공 및 운영
   - 학습자 관리 및 교육 상담
   - 교육 성과 분석 및 통계 작성
   - 안전사고 발생 시 비상연락

2. 수집하는 개인정보의 항목
   - 필수항목: 성명, 생년월일, 성별, 연락처(휴대폰, 자택전화), 주소
   - 비상연락처: 성명, 관계, 연락처
   - 건강상태 및 기타 교육에 필요한 정보

3. 개인정보의 보유·이용 기간
   - 수집·이용 동의일로부터 교육과정 종료 후 5년까지
   - 단, 관계법령에 따른 보존의무가 있는 경우 해당 기간까지 보관

4. 동의를 거부할 권리가 있다는 사실 및 동의 거부에 따른 불이익
   - 개인정보 수집·이용에 대한 동의를 거부하실 수 있습니다.
   - 다만, 동의하지 않으실 경우 교육 프로그램 참여가 제한될 수 있습니다.

위 개인정보의 수집·이용에 동의하십니까?

동의일: {date}
동의자: {name}
서명: [서명란]
"""
    
    if request.method == 'GET':
        if existing_consent:
            form.consent_content.data = existing_consent.consent_content
            form.consent_date.data = existing_consent.consent_date
            form.consent_name.data = existing_consent.consent_name
            form.signature_data.data = existing_consent.signature_data
            form.is_agreed.data = existing_consent.is_agreed
        else:
            # 기본 내용으로 폼 채우기
            form.consent_content.data = default_content.format(
                date=date.today().strftime('%Y년 %m월 %d일'),
                name=learner.name
            )
            form.consent_name.data = learner.name
    
    if form.validate_on_submit():
        if existing_consent:
            # 기존 동의서 업데이트
            existing_consent.consent_content = form.consent_content.data
            existing_consent.consent_date = form.consent_date.data
            existing_consent.consent_name = form.consent_name.data
            existing_consent.signature_data = form.signature_data.data
            existing_consent.is_agreed = form.is_agreed.data
            existing_consent.updated_at = datetime.utcnow()
        else:
            # 새 동의서 생성
            consent = PrivacyConsent(
                learner_id=learner_id,
                consent_content=form.consent_content.data,
                consent_date=form.consent_date.data,
                consent_name=form.consent_name.data,
                signature_data=form.signature_data.data,
                is_agreed=form.is_agreed.data
            )
            db.session.add(consent)
        
        try:
            db.session.commit()
            flash('개인정보 동의서가 저장되었습니다.', 'success')
            return redirect(url_for('learner_profile', learner_id=learner_id))
        except Exception as e:
            db.session.rollback()
            flash('저장 중 오류가 발생했습니다.', 'error')
    
    return render_template('privacy_consent.html', 
                         learner=learner, 
                         form=form, 
                         existing_consent=existing_consent)

@app.route('/privacy_consent_pdf/<int:learner_id>')
def privacy_consent_pdf(learner_id):
    """개인정보 동의서 PDF 다운로드"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
    from reportlab.lib.colors import black
    import io
    import base64
    
    learner = Learner.query.get_or_404(learner_id)
    consent = PrivacyConsent.query.filter_by(learner_id=learner_id).first()
    
    if not consent:
        flash('개인정보 동의서가 없습니다.', 'error')
        return redirect(url_for('learner_profile', learner_id=learner_id))
    
    # PDF 생성
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           rightMargin=2*cm, leftMargin=2*cm,
                           topMargin=2*cm, bottomMargin=2*cm)
    
    # 스타일 정의
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # 중앙 정렬
        fontName='Helvetica-Bold'
    )
    
    content_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=12,
        leading=16,
        fontName='Helvetica'
    )
    
    signature_style = ParagraphStyle(
        'SignatureStyle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=20,
        fontName='Helvetica'
    )
    
    # PDF 내용 구성
    story = []
    
    # 제목
    story.append(Paragraph("개인정보 수집 및 이용 동의서", title_style))
    story.append(Spacer(1, 20))
    
    # 동의서 내용
    content_lines = consent.consent_content.split('\n')
    for line in content_lines:
        if line.strip():
            story.append(Paragraph(line.replace('<', '&lt;').replace('>', '&gt;'), content_style))
        else:
            story.append(Spacer(1, 6))
    
    story.append(Spacer(1, 30))
    
    # 서명 정보
    story.append(Paragraph(f"동의일: {consent.consent_date.strftime('%Y년 %m월 %d일')}", signature_style))
    story.append(Paragraph(f"동의자: {consent.consent_name}", signature_style))
    
    # 서명 이미지 추가
    if consent.signature_data:
        try:
            # base64 서명 데이터를 이미지로 변환
            signature_data = consent.signature_data.split(',')[1] if ',' in consent.signature_data else consent.signature_data
            signature_bytes = base64.b64decode(signature_data)
            signature_buffer = io.BytesIO(signature_bytes)
            
            # 서명 이미지 크기 조정
            story.append(Paragraph("서명:", signature_style))
            signature_img = Image(signature_buffer, width=4*cm, height=2*cm)
            story.append(signature_img)
        except Exception as e:
            story.append(Paragraph("서명: [서명 로드 실패]", signature_style))
    else:
        story.append(Paragraph("서명: [서명 없음]", signature_style))
    
    # PDF 생성
    doc.build(story)
    
    # 응답 생성
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="개인정보동의서_{learner.name}_{consent.consent_date.strftime("%Y%m%d")}.pdf"'
    
    return response

@app.route('/export_excel')
def export_excel():
    """모든 데이터를 엑셀로 내보내기"""
    try:
        # 새 엑셀 워크북 생성
        wb = openpyxl.Workbook()
        
        # 기본 시트 제거
        wb.remove(wb.active)
        
        # 1. 학습자 기본 정보 시트
        learners_ws = wb.create_sheet("학습자 기본정보")
        learners_headers = [
            "ID", "이름", "생년월일", "성별", "휴대폰", "집전화", "주소",
            "비상연락처1_이름", "비상연락처1_관계", "비상연락처1_전화",
            "비상연락처2_이름", "비상연락처2_관계", "비상연락처2_전화",
            "건강상태", "건강상세", "수술경험", "수술상세", "학습경험", "학습기관", "학습기간",
            "입학경로", "학습목적", "읽기수준", "쓰기수준", "수학수준", "기타정보",
            "배정반", "반이름", "상태", "학적상태", "메모", "등록일", "AI분석"
        ]
        learners_ws.append(learners_headers)
        
        learners = Learner.query.all()
        for learner in learners:
            row = [
                learner.id, learner.name, 
                learner.birth_date.strftime('%Y-%m-%d') if learner.birth_date else '',
                learner.gender, learner.phone, learner.home_phone or '', learner.address,
                learner.emergency_contact1_name, learner.emergency_contact1_relation, learner.emergency_contact1_phone,
                learner.emergency_contact2_name or '', learner.emergency_contact2_relation or '', learner.emergency_contact2_phone or '',
                learner.health_status, learner.health_details or '', learner.surgery_experience, learner.surgery_details or '',
                learner.learning_experience, learner.learning_institution or '', learner.learning_period or '',
                learner.enrollment_path, learner.learning_purpose, learner.reading_level, learner.writing_level, learner.math_level,
                learner.other_info or '', learner.assigned_class or '', learner.class_name or '', learner.status, learner.academic_status,
                learner.memo or '', learner.created_at.strftime('%Y-%m-%d %H:%M') if learner.created_at else '',
                learner.ai_analysis[:100] + '...' if learner.ai_analysis and len(learner.ai_analysis) > 100 else learner.ai_analysis or ''
            ]
            learners_ws.append(row)
        
        # 2. 레벨 테스트 결과 시트
        leveltest_ws = wb.create_sheet("레벨테스트결과")
        leveltest_headers = [
            "학습자ID", "학습자명", "완료날짜", "총점", "어휘점수", "문법점수", "독해점수",
            "기초답안", "중급답안", "고급답안", "그림데이터", "상세평가"
        ]
        leveltest_ws.append(leveltest_headers)
        
        level_tests = LevelTest.query.join(Learner).all()
        for test in level_tests:
            row = [
                test.learner_id, test.learner.name,
                test.completed_at.strftime('%Y-%m-%d %H:%M') if test.completed_at else '',
                test.total_score, test.vocabulary_score, test.grammar_score, test.reading_score,
                str(test.basic_answers)[:50] + '...' if test.basic_answers else '',
                str(test.intermediate_answers)[:50] + '...' if test.intermediate_answers else '',
                str(test.advanced_answers)[:50] + '...' if test.advanced_answers else '',
                '그림데이터 있음' if test.drawing_data else '없음',
                test.details or ''
            ]
            leveltest_ws.append(row)
        
        # 3. 상담 기록 시트
        counseling_ws = wb.create_sheet("상담기록")
        counseling_headers = [
            "학습자ID", "학습자명", "상담자명", "상담내용", "추천사항", "비고", "작성일", "수정일"
        ]
        counseling_ws.append(counseling_headers)
        
        counselings = Counseling.query.join(Learner).all()
        for counseling in counselings:
            row = [
                counseling.learner_id, counseling.learner.name,
                counseling.counselor_name, counseling.counseling_content,
                counseling.recommendations or '', counseling.notes or '',
                counseling.created_at.strftime('%Y-%m-%d %H:%M') if counseling.created_at else '',
                counseling.updated_at.strftime('%Y-%m-%d %H:%M') if counseling.updated_at else ''
            ]
            counseling_ws.append(row)
        
        # 4. 시험 기록 시트
        exams_ws = wb.create_sheet("시험기록")
        exams_headers = [
            "학습자ID", "학습자명", "시험명", "시험날짜", "점수", "만점", "비고"
        ]
        exams_ws.append(exams_headers)
        
        exams = Exam.query.join(Learner).all()
        for exam in exams:
            row = [
                exam.learner_id, exam.learner.name, exam.exam_name,
                exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else '',
                exam.score, exam.max_score, exam.notes
            ]
            exams_ws.append(row)
        
        # 5. 상담 접수 기록 시트
        consultation_ws = wb.create_sheet("상담접수기록")
        consultation_headers = [
            "ID", "이름", "전화번호", "상담내용", "상담날짜", "성별", "거주지", "상담영역", "접수일"
        ]
        consultation_ws.append(consultation_headers)
        
        consultations = ConsultationRecord.query.all()
        for consultation in consultations:
            row = [
                consultation.id, consultation.name or '', consultation.phone,
                consultation.consultation_content or '',
                consultation.consultation_date.strftime('%Y-%m-%d') if consultation.consultation_date else '',
                consultation.gender or '', consultation.residence or '', consultation.consultation_area or '',
                consultation.created_at.strftime('%Y-%m-%d %H:%M') if consultation.created_at else ''
            ]
            consultation_ws.append(row)
        
        # 6. 학습자 이력 시트
        history_ws = wb.create_sheet("학습자이력")
        history_headers = [
            "학습자ID", "학습자명", "이벤트유형", "이벤트날짜", "이벤트제목",
            "이벤트설명", "이전값", "새값", "기록일", "기록자"
        ]
        history_ws.append(history_headers)
        
        histories = LearnerHistory.query.join(Learner).all()
        for history in histories:
            row = [
                history.learner_id, history.learner.name, history.event_type,
                history.event_date.strftime('%Y-%m-%d') if history.event_date else '',
                history.event_title, history.event_description,
                history.previous_value, history.new_value,
                history.created_at.strftime('%Y-%m-%d %H:%M') if history.created_at else '',
                history.created_by
            ]
            history_ws.append(row)
        
        # 7. 개인정보 동의 시트
        privacy_ws = wb.create_sheet("개인정보동의")
        privacy_headers = [
            "학습자ID", "학습자명", "동의날짜", "동의자명", "동의여부", "작성일", "수정일"
        ]
        privacy_ws.append(privacy_headers)
        
        privacy_consents = PrivacyConsent.query.join(Learner).all()
        for consent in privacy_consents:
            row = [
                consent.learner_id, consent.learner.name,
                consent.consent_date.strftime('%Y-%m-%d') if consent.consent_date else '',
                consent.consent_name, "동의" if consent.is_agreed else "미동의",
                consent.created_at.strftime('%Y-%m-%d %H:%M') if consent.created_at else '',
                consent.updated_at.strftime('%Y-%m-%d %H:%M') if consent.updated_at else ''
            ]
            privacy_ws.append(row)
        
        # 엑셀 스타일링
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 모든 시트에 스타일 적용
        for ws in wb.worksheets:
            # 헤더 행 스타일 적용
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 메모리에서 엑셀 파일 생성
        from io import BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 현재 날짜로 파일명 생성
        from datetime import datetime
        current_date = datetime.now().strftime('%Y%m%d')
        filename = f"난곡사랑의집_전체데이터_{current_date}.xlsx"
        
        # 응답 생성
        response = make_response(excel_buffer.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        flash(f'엑셀 내보내기 중 오류가 발생했습니다: {str(e)}', 'error')
        return redirect(url_for('upload'))

# 레벨 테스트에서 동적 문제 사용
def get_level_test_questions(level):
    """관리자가 설정한 문제들을 동적으로 로드"""
    from admin_models import TestQuestion
    questions = TestQuestion.query.filter_by(level=level, is_active=True).order_by(TestQuestion.order_index).all()
    
    # 기존 하드코딩된 문제를 백업으로 유지
    if not questions:
        # 기존 LEVEL_TEST_QUESTIONS 사용
        return LEVEL_TEST_QUESTIONS.get(level, [])
    
    return [q.to_dict() for q in questions]

@app.route('/transcribe-audio', methods=['POST'])
def transcribe_audio_route():
    """음성 파일을 텍스트로 변환"""
    try:
        if 'audio' not in request.files:
            return jsonify({'success': False, 'error': '음성 파일이 없습니다.'})
        
        audio_file = request.files['audio']
        if audio_file.filename == '':
            return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다.'})
        
        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix='.wav') as tmp_file:
            audio_file.save(tmp_file.name)
            
            # 음성을 텍스트로 변환
            transcription = transcribe_audio(tmp_file.name)
            
            # 임시 파일 삭제
            os.unlink(tmp_file.name)
            
            if transcription:
                return jsonify({'success': True, 'text': transcription})
            else:
                return jsonify({'success': False, 'error': '음성 변환에 실패했습니다.'})
                
    except Exception as e:
        return jsonify({'success': False, 'error': f'서버 오류: {str(e)}'})

@app.route('/voice-to-summary', methods=['POST'])
def voice_to_summary_route():
    """음성을 바로 AI 요약으로 변환"""
    try:
        if 'audio' not in request.files:
            return jsonify({'success': False, 'error': '음성 파일이 없습니다.'})
        
        audio_file = request.files['audio']
        learner_id = request.form.get('learner_id')
        
        if audio_file.filename == '':
            return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다.'})
        
        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix='.wav') as tmp_file:
            audio_file.save(tmp_file.name)
            
            # 음성을 텍스트로 변환
            transcription = transcribe_audio(tmp_file.name)
            
            # 임시 파일 삭제
            os.unlink(tmp_file.name)
            
            if not transcription:
                return jsonify({'success': False, 'error': '음성 변환에 실패했습니다.'})
            
            # 학습자 정보 가져오기
            learner_info = None
            if learner_id:
                learner = Learner.query.get(learner_id)
                if learner:
                    learner_info = {
                        'name': learner.name,
                        'birth_date': learner.birth_date.strftime('%Y년 %m월 %d일') if learner.birth_date else '',
                        'learning_experience': learner.learning_experience,
                        'learning_purpose': learner.learning_purpose
                    }
            
            # AI 요약 실행
            summary = summarize_counseling_session(transcription, learner_info)
            
            if summary:
                return jsonify({'success': True, 'summary': summary, 'transcription': transcription})
            else:
                return jsonify({'success': False, 'error': 'AI 요약에 실패했습니다.'})
                
    except Exception as e:
        return jsonify({'success': False, 'error': f'서버 오류: {str(e)}'})

@app.route('/summarize-counseling', methods=['POST'])
def summarize_counseling_route():
    """상담 내용을 AI로 요약"""
    try:
        data = request.get_json()
        content = data.get('content', '')
        learner_id = data.get('learner_id')
        
        if not content.strip():
            return jsonify({'success': False, 'error': '요약할 내용이 없습니다.'})
        
        # 학습자 정보 가져오기
        learner_info = None
        if learner_id:
            learner = Learner.query.get(learner_id)
            if learner:
                learner_info = {
                    'name': learner.name,
                    'birth_date': learner.birth_date.strftime('%Y년 %m월 %d일') if learner.birth_date else '',
                    'learning_experience': learner.learning_experience,
                    'learning_purpose': learner.learning_purpose
                }
        
        # AI 요약 실행
        summary = summarize_counseling_session(content, learner_info)
        
        if summary:
            return jsonify({'success': True, 'summary': summary})
        else:
            return jsonify({'success': False, 'error': 'AI 요약에 실패했습니다.'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': f'서버 오류: {str(e)}'})


@app.route("/personal_story/<int:learner_id>")
def personal_story(learner_id):
    """개인 상황 이야기 페이지"""
    learner = Learner.query.get_or_404(learner_id)
    
    # 개인 이야기 기록 조회 (아직 테이블이 없으므로 빈 리스트)
    personal_stories = []
    
    return render_template("personal_story.html", 
                         learner=learner, 
                         personal_stories=personal_stories)

@app.route("/submit_personal_story", methods=["POST"])
def submit_personal_story():
    """개인 상황 이야기 제출 및 AI 요약"""
    try:
        data = request.get_json()
        learner_id = data.get("learner_id")
        audio_data = data.get("audio_data")
        
        if not learner_id or not audio_data:
            return jsonify({"error": "필수 데이터가 누락되었습니다."}), 400
        
        learner = Learner.query.get_or_404(learner_id)
        
        # 임시로 학습자의 메모에 개인 이야기 정보 추가
        current_memo = learner.memo or ""
        timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
        story_entry = f"\n[개인 이야기 - {timestamp}]\n음성 녹음 완료\n"
        learner.memo = current_memo + story_entry
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "개인 상황 이야기가 성공적으로 저장되었습니다."
        })
        
    except Exception as e:
        print(f"개인 이야기 저장 오류: {e}")
        return jsonify({"error": "저장 중 오류가 발생했습니다."}), 500

